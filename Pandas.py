import os
import pandas as pd
import sqlite3
from openpyxl import load_workbook
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill

class PandasObject:
    def __init__(self, filename):
        self.filename = filename

    def get_excel_row_count(self, sheet_name):
        wb = load_workbook(self.filename, read_only=True)
        sheet = wb[sheet_name]
        row_count = sheet.max_row
        wb.close()
        return row_count

    def get_root_path(self):
        return os.path.dirname(os.path.abspath(__file__))

    def read_excel_data(self, sheet_name, header_rows, columns_range):
        excel_file_path = os.path.join(self.get_root_path(), self.filename)

        data_rows = range(0, self.get_excel_row_count(sheet_name) + 1)

        # Read data from Excel
        df = pd.read_excel(excel_file_path,
                           sheet_name,
                           header=header_rows,
                           usecols=columns_range,
                           skiprows=lambda x: x not in data_rows)
        return df

    def write_to_database(self, df, databaseName, sheet_name):
        if sheet_name == 'Sheet1': # Change sheet_name according to your preference
            # Add a boolean column 'isArchive'
            df['isArchive'] = False  # Initialize all rows with False

            # Find the rows below the "REMOVED" and set 'isArchive' to True
            remove_index = df[df['Title'] == 'REMOVED'].index[0] # Change this part according to your preference
            df.loc[remove_index + 1:, 'isArchive'] = True

            # Filter out 'REMOVED'
            df = df[df['Title'] != 'REMOVED']

            # Filter out rows with null values 
            df = df[df['Reference'].notnull()]

        databaseName = f"{databaseName}.db"

        db_path = os.path.join(self.get_root_path(), databaseName)

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        table_name = 'my_data_table'

        # Save the DataFrame to SQLite, making 'Reference' the primary key
        df.to_sql(table_name, conn, index=False, if_exists='replace')

        conn.close()

        print(f"Data imported from Excel to SQLite table '{table_name}'.")
        print("\n")
        


    def compare_differences(self, primary_db, secondary_db):
        print(f"Comparing databases '{primary_db}' and '{secondary_db}'...")
        primary_conn = sqlite3.connect(primary_db)
        secondary_conn = sqlite3.connect(secondary_db)

        primary_df = pd.read_sql_query("SELECT * FROM my_data_table", primary_conn)
        secondary_df = pd.read_sql_query("SELECT * FROM my_data_table", secondary_conn)

        primary_conn.close()
        secondary_conn.close()

        # Change this part according to your preference
        differences_df = pd.DataFrame(
            columns=['Reference', '\nIssue', '\nValidation Status', 'Confidence Level', '\nValidation\nStatus.1',
                     'Confidence Level.2'])

        common_references_and_issues = set(zip(primary_df['Reference'], primary_df['\nIssue'])) & set(zip(secondary_df['Reference'], secondary_df['\nIssue']))

        # highlight cells with "X"
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for reference, issue in common_references_and_issues:
            is_different = False

            primary_row = primary_df[(primary_df['Reference'] == reference) & (primary_df['\nIssue'] == issue)]
            secondary_row = secondary_df[(secondary_df['Reference'] == reference) & (secondary_df['\nIssue'] == issue)]
            
            if 'isArchive' in primary_row.columns and primary_row['isArchive'].iloc[0] == 1:
                continue

            columns_to_check = ['\nValidation Status', 'Confidence Level', '\nValidation\nStatus.1', 'Confidence Level.2']

            for col in columns_to_check:
                primary = primary_row[col].iloc[0]
                secondary = secondary_row[col].iloc[0]

                if primary != secondary:
                    primary_row[col].iloc[0] = "X"
                    is_different = True
                    # highlight
                    primary_row[col].fill = fill

            if is_different:
                differences_df = pd.concat([differences_df, primary_row], ignore_index=True)

        if not differences_df.empty:
            # Filter the table to include only rows where these columns have X values
            filtered_df = differences_df[(differences_df['\nValidation Status'] == 'X') | (differences_df['Confidence Level'] == 'X') | (differences_df['\nValidation\nStatus.1'] == 'X') | (differences_df['Confidence Level.2'] == 'X')]

            excel_output_path = os.path.join(self.get_root_path(), "differences.xlsx")

            # Rename headers
            filtered_df = filtered_df.rename(columns={
                'Reference': 'REFERENCE',
                '\nIssue': 'ISSUE',
                '\nValidation Status': 'UT VAL RESULT',
                'Confidence Level': 'UT CONF LVL',
                '\nValidation\nStatus.1': 'FT VAL RESULT',
                'Confidence Level.2': 'FT CONF LVL'
            })

            filtered_df.to_excel(excel_output_path, index=False,
                                 columns=['REFERENCE', 'ISSUE', 'UT VAL RESULT', 'UT CONF LVL',
                                          'FT VAL RESULT', 'FT CONF LVL'])

            # Load the Excel file and apply the fill
            wb = openpyxl.load_workbook(excel_output_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=8):
                for cell in row:
                    if cell.value == "X":
                        cell.fill = fill
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)  # text wrapping
                        # Set column width to accommodate the content
                        sheet.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width = 15  # adjust the width 

            # Adjust column widths for headers
            for col_num, column_header in enumerate(filtered_df.columns, 1):
                max_length = max(filtered_df[column_header].astype(str).apply(len).max(), len(column_header))
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max_length + 2 

            wb.save(excel_output_path)

            print("Differences saved to Excel file 'differences.xlsx'.")

